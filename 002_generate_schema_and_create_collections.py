#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
002 - 生成分类Schema并创建Zotero集合
基于LLM分析文献生成分类体系，创建Zotero集合结构

主要功能：
1. 使用LLM分析文献生成分类体系
2. 修复LLM生成的schema格式
3. 创建Zotero集合结构
4. 支持dry-run和测试模式

注意：此脚本仅负责schema生成和集合创建，文献分类请使用006脚本
"""

import os
import sys
import json
import argparse
from pathlib import Path
from typing import Dict, Any, Optional, List
import pandas as pd
import requests
from datetime import datetime

# 导入配置系统
from config import (
    get_llm_config, get_zotero_config, get_config,
    get_default_max_items, get_default_test_items, get_default_dry_run_items,
    get_max_tokens_limit, get_default_output_tokens, get_description_preview_length
)
from llm_client import LLMClient

# 导入自定义模块（延迟导入）
# from llm_client import LLMClient

# 设置日志
import logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

from pydantic import BaseModel, ValidationError, Field
import re

class SubCategoryModel(BaseModel):
    name: str
    description: str

class MainCategoryModel(BaseModel):
    name: str
    description: str
    subcategories: dict[str, SubCategoryModel]

class ClassificationSchemaModel(BaseModel):
    main_categories: dict[str, MainCategoryModel]


def verify_schema(schema: dict) -> list[str]:
    """验证LLM生成的schema结构和内容是否合法，返回错误列表"""
    errors = []
    try:
        model = ClassificationSchemaModel(**schema)
    except ValidationError as e:
        errors.append(f"结构校验失败: {e}")
        return errors
    
    main_categories = schema.get("main_categories", {})
    if not (5 <= len(main_categories) <= 20):
        errors.append(f"主分类数量不在5-20范围: {len(main_categories)}")
    
    for code, main_cat in main_categories.items():
        name = main_cat.get("name", "")
        
        # 主分类名称验证
        if not name.startswith("[AUTO]"):
            errors.append(f"主分类 {code} 名称未以[AUTO]开头: {name}")
        
        # 主分类词数验证（移除[AUTO]前缀后计算）
        clean_name = name.replace('[AUTO]', '').strip()
        word_count = len(clean_name.split())
        if not (1 <= word_count <= 10):
            errors.append(f"主分类 {code} 名称词数不在1-10: {name}")
        
        # 子分类验证
        subcats = main_cat.get("subcategories", {})
        if not (2 <= len(subcats) <= 10):
            errors.append(f"主分类 {code} 子分类数量不在2-10: {len(subcats)}")
        
        for sub_code, sub_cat in subcats.items():
            sub_name = sub_cat.get("name", "")
            
            # 子分类词数验证（更宽松，允许1-10个词）
            sub_word_count = len(sub_name.split())
            if not (1 <= sub_word_count <= 10):
                errors.append(f"子分类 {sub_code} 名称词数不在1-10: {sub_name}")
    
    return errors

class SchemaBasedCollectionManager:
    """基于schema的集合管理器"""
    
    def __init__(self, init_llm: bool = True, init_zotero: bool = True):
        """初始化管理器"""
        # 可选初始化LLM和Zotero客户端
        self.llm_client = self._init_llm_client() if init_llm else None
        self.zotero_client = self._init_zotero_client() if init_zotero else None
        
        self.collection_keys = {}
        
        # 统计信息
        self.collections_created = 0
    
    def _init_llm_client(self):
        """初始化LLM客户端"""
        try:
            from llm_client import LLMClient
        except ImportError as e:
            logger.error(f"❌ 无法导入LLMClient: {e}")
            return None
        
        # 使用新的配置系统
        return LLMClient()
    
    def _init_zotero_client(self):
        """初始化Zotero客户端"""
        # 使用新的配置系统
        zotero_config = get_zotero_config()
        
        return {
            'user_id': zotero_config.user_id,
            'api_key': zotero_config.api_key,
            'base_url': zotero_config.api_base_url,
            'headers': zotero_config.headers
        }
    
    def _estimate_tokens(self, text: str) -> int:
        """估算token数量（改进版：支持中英文混合文本）"""
        if not text:
            return 0
        
        # 分离中文字符和英文单词
        import re
        
        # 中文字符（包括中文标点）
        chinese_chars = re.findall(r'[\u4e00-\u9fff\u3000-\u303f\uff00-\uffef]', text)
        chinese_token_count = len(chinese_chars)
        
        # 英文单词和其他字符
        non_chinese_text = re.sub(r'[\u4e00-\u9fff\u3000-\u303f\uff00-\uffef]', ' ', text)
        english_words = non_chinese_text.split()
        english_token_count = len(english_words) * 1.3  # 英文单词按1.3倍计算
        
        # 其他字符（数字、标点等）
        other_chars = re.findall(r'[^\u4e00-\u9fff\u3000-\u303f\uff00-\uffef\s\w]', text)
        other_token_count = len(other_chars) * 0.5  # 其他字符按0.5倍计算
        
        total_tokens = chinese_token_count + english_token_count + other_token_count
        return int(total_tokens)
    
    def generate_collections_from_literature(self, literature_file: str, max_items: int = None, dry_run: bool = False, return_schema_only: bool = False) -> Dict[str, str]:
        # 使用配置系统获取默认值
        if max_items is None:
            max_items = get_default_max_items()
        """使用LLM分析文献生成合理的集合分类"""
        if dry_run:
            logger.info("🔍 DRY RUN模式：展示LLM生成集合的计划...")
        else:
            logger.info("🧠 使用LLM分析文献生成集合分类...")
        
        # 加载文献数据
        try:
            df = pd.read_excel(literature_file)
            logger.info(f"✅ 成功加载文献数据: {len(df)} 篇文献")
        except Exception as e:
            logger.error(f"❌ 加载文献数据失败: {e}")
            return {}
        
        # 使用所有文献进行分析
        logger.info(f"📊 将分析所有 {len(df)} 篇文献来生成分类")
        
        # 获取现有集合（仅在非测试模式下）
        existing_collections = {}
        if not return_schema_only:
            existing_collections = self._get_existing_collections()
        
        if dry_run:
            print(f"\n📊 LLM生成集合计划:")
            print(f"分析文献数: {len(df)} 篇")
            print(f"现有集合: {len(existing_collections)} 个")
            print(f"预计LLM调用: 1次（生成分类体系）")
            print("\n📋 分析内容:")
            print(f"- 文献标题和摘要")
            print(f"- 研究领域分布")
            print(f"- 主题聚类分析")
            print(f"- 生成层次化分类体系")
            return {}
        
        # 准备所有文献样本用于LLM分析（使用更长的摘要）
        literature_samples = []
        for idx, row in df.iterrows():
            title = str(row.get('title', '')).strip()
            abstract = str(row.get('abstract', '')).strip()
            
            if title and abstract:
                # 保留完整摘要，不截断，让LLM获得更多信息
                literature_samples.append({
                    'title': title,
                    'abstract': abstract
                })
        
        if not literature_samples:
            logger.error("❌ 没有找到有效的文献样本")
            return {}
        
        # 创建LLM提示词来生成分类体系
        system_prompt = """You are a professional academic literature classification expert specializing in Computer Science and related fields. Your task is to design a comprehensive and well-balanced classification system based on the provided literature samples.\n\nPlease carefully analyze ALL titles and abstracts in the literature samples to identify the complete spectrum of research areas, technical topics, and disciplinary directions. Pay special attention to:\n\n1. **Comprehensive Coverage**: Ensure the classification covers ALL major research areas present in the literature, including but not limited to:\n   - AI/ML (Foundation Models, LLMs, MLLMs, Computer Vision, etc.)\n   - Traditional Systems (Distributed Systems, Operating Systems, Database Systems, etc.)\n   - AI Systems (Training Frameworks, Inference Frameworks, GPU Optimaztions, Attention Optimaztions, etc.)\n   - Scientific Computing (HPC, Physics, Chemistry, Biology applications, etc.)\n   - Graphics and Visualization (3D rendering, 3DGS, NeRF, Computer Graphics, etc.)\n   - Programming Languages and Software Engineering\n   - Infras (Networks, Storage, Security, DataCenters etc.)\n   - Any other CS domains present in the literature\n\n2. **Balanced Representation**: Ensure that all significant research areas in the literature are represented proportionally, without over-emphasizing any single domain.\n\n3. **Hierarchical Structure**: Create a logical two-level hierarchy where:\n   - Main categories represent broad CS research domains\n   - Subcategories represent specific technical directions within each domain\n\n4. **Professional Standards**: Use standard CS terminology and naming conventions that would be recognized by the academic community.\n\nClassification System Requirements:\n1. Main Categories: 8-15 major Computer Science research domains\n2. Subcategories: Each main category must have 4-10 specific technical directions (aim for more granular subcategories)\n3. Category Names: Concise, professional English names using 2-4 words maximum (never exceed 5 words), using standard CS terminology. IMPORTANT: All main category names must be prefixed with "[AUTO]" (e.g., "[AUTO] AI and Machine Learning Models")\n4. Category Descriptions: Clear, accurate descriptions of research scope and content\n5. Coverage: Ensure ALL significant research areas from the literature are covered\n6. Balance: Avoid over-representing any single domain while ensuring comprehensive coverage\n\nPlease return the classification system in JSON format:"""

        # 构建完整的文献文本（使用所有文献）
        literature_text = ""
        for i, sample in enumerate(literature_samples):
            literature_text += f"{i+1}. 标题：{sample['title']}\n   摘要：{sample['abstract']}\n\n"

        user_prompt = f"""Please design a comprehensive and well-balanced Computer Science classification system based on the following {len(literature_samples)} literature samples.\n\nLiterature Samples:\n{literature_text}\n\nIMPORTANT: Please ensure that your classification system:\n1. Covers ALL significant research areas present in the literature samples\n2. Provides balanced representation across different CS domains\n3. Includes specific categories for any specialized research areas (e.g., High Energy Physics, Scientific Computing, etc.)\n4. Uses standard Computer Science terminology and naming conventions\n5. Category names should be concise (2-4 words, maximum 5 words)\n6. Create more granular subcategories (4-10 per main category) for better organization\n7. ALL main category names MUST be prefixed with "[AUTO]" (e.g., "[AUTO] AI and Machine Learning Models")\n\nPlease return the classification system in JSON format:\n{{\n    "main_categories": {{\n        "category_code": {{\n            "name": "Category Name",\n            "description": "Category Description",\n            "subcategories": {{\n                "subcategory_code": {{\n                    "name": "Subcategory Name", \n                    "description": "Subcategory Description"\n                }}\n            }}\n        }}\n    }}\n}}\n\nRequirements:\n- All category names and descriptions must be in English\n- Use professional Computer Science terminology\n- Ensure comprehensive coverage of ALL research areas in the literature\n- Maintain clear hierarchical structure with main categories and subcategories\n- Focus on Computer Science domains while including interdisciplinary applications\n- Category names must be concise (2-4 words, never exceed 5 words)\n- Create more granular subcategories (3-6 per main category) for better organization\n- ALL main category names MUST be prefixed with "[AUTO]" (e.g., "[AUTO] AI and Machine Learning Models")"""
        
        # 计算token使用量
        total_prompt = system_prompt + "\n\n" + user_prompt
        estimated_tokens = self._estimate_tokens(total_prompt)
        # 简单的成本估算（美元）
        input_cost = (estimated_tokens / 1000) * 0.0035
        default_output_tokens = get_default_output_tokens()
        output_cost = (default_output_tokens / 1000) * 0.105
        estimated_cost = input_cost + output_cost
        
        logger.info(f"📊 Token估算:")
        logger.info(f"  分析文献数: {len(literature_samples)} 篇")
        logger.info(f"  输入tokens: ~{estimated_tokens:,}")
        logger.info(f"  输出tokens: ~{default_output_tokens:,}")
        logger.info(f"  总tokens: ~{estimated_tokens + default_output_tokens:,}")
        logger.info(f"  估算成本: ${estimated_cost:.4f}")
        
        # 检查token限制
        max_tokens_limit = get_max_tokens_limit()
        if estimated_tokens > max_tokens_limit:
            logger.warning(f"⚠️  估算token数量 ({estimated_tokens:,}) 超过{max_tokens_limit:,}限制")
            logger.info("建议减少文献数量或截断摘要")
            return {}
        
        # 用户确认（仅在非dry_run且非return_schema_only时）
        if not dry_run and not return_schema_only:
            print(f"\n📊 LLM请求估算:")
            print(f"  分析文献数: {len(literature_samples)} 篇")
            print(f"  估算输入tokens: {estimated_tokens:,}")
            print(f"  估算输出tokens: {default_output_tokens:,}")
            print(f"  总tokens: {estimated_tokens + default_output_tokens:,}")
            print(f"  估算成本: ${estimated_cost:.4f}")
            
            confirm = input("\n是否继续执行？(y/N): ").strip().lower()
            if confirm != 'y':
                logger.info("用户取消操作")
                return {}

        # 调用LLM生成分类体系
        try:
            response = self.llm_client.generate(
                prompt=user_prompt,
                system_prompt=system_prompt,
                max_tokens=default_output_tokens,
                temperature=0.3
            )
            # 解析LLM响应
            classification_system = self._parse_classification_system(response.get('content', ''))
            if not classification_system:
                logger.error("❌ LLM生成的分类体系解析失败")
                return {}
            # 在LLM生成schema后调用verify_schema
            errors = verify_schema(classification_system)
            if errors:
                logger.warning("⚠️ LLM生成的schema未通过校验：")
                for err in errors:
                    logger.warning(f"   - {err}")
            else:
                logger.info("✅ LLM成功生成分类体系")
            # 如果只需要返回分类体系，不创建集合
            if return_schema_only:
                return classification_system
            # 创建集合
            return self._create_collections_from_llm_system(classification_system, existing_collections, dry_run)
        except Exception as e:
            logger.error(f"❌ LLM生成分类体系失败: {e}")
            return {}
    
    def _parse_classification_system(self, response: str) -> Dict[str, Any]:
        """解析LLM生成的分类体系"""
        try:
            # 尝试提取JSON部分
            start_idx = response.find('{')
            end_idx = response.rfind('}') + 1
            
            if start_idx == -1 or end_idx == 0:
                logger.error("无法找到JSON格式的分类体系")
                return {}
            
            json_str = response[start_idx:end_idx]
            classification_system = json.loads(json_str)
            
            # 验证结构
            if 'main_categories' not in classification_system:
                logger.error("分类体系缺少main_categories字段")
                return {}
            
            return classification_system
            
        except json.JSONDecodeError as e:
            logger.error(f"JSON解析失败: {e}")
            return {}
        except Exception as e:
            logger.error(f"解析分类体系失败: {e}")
            return {}
    
    def save_llm_generated_schema(self, classification_system: Dict[str, Any], output_file: str = None) -> str:
        """保存LLM生成的原始schema"""
        if output_file is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = f"data/llm_generated_schema_{timestamp}.json"
        
        try:
            os.makedirs(os.path.dirname(output_file), exist_ok=True)
            with open(output_file, 'w', encoding='utf-8') as f:
                json.dump(classification_system, f, ensure_ascii=False, indent=2)
            
            logger.info(f"✅ LLM生成的schema已保存到: {output_file}")
            return output_file
        except Exception as e:
            logger.error(f"❌ 保存LLM生成的schema失败: {e}")
            return ""
    
    def save_ready_schema(self, classification_system: Dict[str, Any]) -> str:
        """保存为ready状态的schema文件（第一步输出）"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = f"data/classification_schema_ready_{timestamp}.json"
        excel_output_file = f"data/classification_schema_ready_{timestamp}.xlsx"
        
        # 计算统计信息
        main_categories = classification_system.get('main_categories', {})
        total_main = len(main_categories)
        total_sub = sum(len(cat.get('subcategories', {})) for cat in main_categories.values())
        
        # 生成预览信息
        preview = self._generate_schema_preview(classification_system)
        
        # 构建ready状态的schema
        ready_schema = {
            "metadata": {
                "generated_at": datetime.now().isoformat(),
                "status": "ready_for_review",
                "total_main_categories": total_main,
                "total_sub_categories": total_sub,
                "total_categories": total_main + total_sub
            },
            "classification_schema": classification_system,
            "preview": preview
        }
        
        try:
            os.makedirs(os.path.dirname(output_file), exist_ok=True)
            with open(output_file, 'w', encoding='utf-8') as f:
                json.dump(ready_schema, f, ensure_ascii=False, indent=2)
            
            logger.info(f"✅ Ready schema已保存到: {output_file}")
            
            # 保存Excel文件
            self._save_schema_to_excel(classification_system, excel_output_file)
            
            return output_file
        except Exception as e:
            logger.error(f"❌ 保存ready schema失败: {e}")
            return ""

    def _save_schema_to_excel(self, classification_system: Dict[str, Any], excel_output_file: str):
        """将生成的schema保存到Excel文件"""
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            excel_data = []
            main_categories = classification_system.get('main_categories', {})

            for main_cat_code, main_cat_info in main_categories.items():
                subcategories = main_cat_info.get('subcategories', {})
                if subcategories:
                    for sub_cat_code, sub_cat_info in subcategories.items():
                        row_data = {
                            'main_category_code': main_cat_code,
                            'main_category_name': main_cat_info.get('name', ''),
                            'main_category_description': main_cat_info.get('description', ''),
                            'subcategory_code': sub_cat_code,
                            'subcategory_name': sub_cat_info.get('name', ''),
                            'subcategory_description': sub_cat_info.get('description', '')
                        }
                        excel_data.append(row_data)
                else:
                    row_data = {
                        'main_category_code': main_cat_code,
                        'main_category_name': main_cat_info.get('name', ''),
                        'main_category_description': main_cat_info.get('description', ''),
                        'subcategory_code': '',
                        'subcategory_name': '',
                        'subcategory_description': ''
                    }
                    excel_data.append(row_data)

            df = pd.DataFrame(excel_data)

            with pd.ExcelWriter(excel_output_file, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Classification Schema', index=False)
                worksheet = writer.sheets['Classification Schema']

                # 设置列宽
                column_widths = {
                    'A': 20,
                    'B': 40,
                    'C': 60,
                    'D': 20,
                    'E': 40,
                    'F': 60,
                }
                for col, width in column_widths.items():
                    worksheet.column_dimensions[col].width = width

                # 设置标题行样式
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")

                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment

                # 设置数据行样式
                for row in worksheet.iter_rows(min_row=2):
                    for cell in row:
                        cell.alignment = Alignment(wrap_text=True, vertical="top")

            logger.info(f"✅ Schema已导出到Excel文件: {excel_output_file}")

        except ImportError as e:
            logger.warning(f"⚠️ 无法生成Excel文件，缺少依赖: {e}")
            logger.info("请安装: pip install pandas openpyxl")
        except Exception as e:
            logger.error(f"❌ 生成Excel文件失败: {e}")
    
    def _generate_schema_preview(self, classification_system: Dict[str, Any]) -> Dict[str, Any]:
        """生成schema预览信息"""
        main_categories = classification_system.get('main_categories', {})
        
        preview = {
            "summary": f"基于文献分析，生成了{len(main_categories)}个主分类",
            "main_categories": [],
            "recommendations": [
                "建议检查分类名称是否符合Zotero命名规范",
                "确认分类结构是否合理",
                "检查是否有重复或过于细分的分类"
            ]
        }
        
        for cat_name, cat_info in main_categories.items():
            sub_cats = cat_info.get('subcategories', {})
            preview["main_categories"].append({
                "name": cat_name,
                "description": cat_info.get('description', ''),
                "sub_categories_count": len(sub_cats),
                "sub_categories": list(sub_cats.keys())
            })
        
        return preview
    
    def get_operation_summary(self, schema_file: str) -> Dict[str, Any]:
        """获取操作摘要信息"""
        try:
            with open(schema_file, 'r', encoding='utf-8') as f:
                schema_data = json.load(f)
            
            classification_system = schema_data.get('classification_schema', {})
            main_categories = classification_system.get('main_categories', {})
            
            total_main = len(main_categories)
            total_sub = sum(len(cat.get('subcategories', {})) for cat in main_categories.values())
            
            return {
                'main_categories': total_main,
                'sub_categories': total_sub,
                'total_categories': total_main + total_sub
            }
        except Exception as e:
            logger.error(f"❌ 读取schema文件失败: {e}")
            return {'main_categories': 0, 'sub_categories': 0, 'total_categories': 0}
    
    def create_collections_from_ready_schema(self, schema_file: str, dry_run: bool = False) -> str:
        """从ready schema创建集合（第二步）"""
        try:
            # 读取ready schema
            with open(schema_file, 'r', encoding='utf-8') as f:
                schema_data = json.load(f)
            
            classification_system = schema_data.get('classification_schema', {})
            if not classification_system:
                logger.error("❌ Schema文件中没有找到分类体系")
                return ""
            
            # 获取现有集合
            existing_collections = self._get_existing_collections()
            
            # 创建集合
            collection_keys = self._create_collections_from_llm_system(
                classification_system, 
                existing_collections, 
                dry_run=dry_run
            )
            
            if not collection_keys:
                logger.error("❌ 创建集合失败")
                return ""
            
            # 保存带collection keys的完整schema
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = f"data/schema_with_collection_keys_{timestamp}.json"
            
            # 构建完整schema
            complete_schema = {
            "metadata": {
                    "created_at": datetime.now().isoformat(),
                    "status": "collections_created",
                    "source_file": schema_file,
                    "total_collections_created": len(collection_keys)
                },
                "classification_schema": classification_system,
                "collection_mapping": collection_keys
            }
            try:
                with open(output_file, 'w', encoding='utf-8') as f:
                    json.dump(complete_schema, f, ensure_ascii=False, indent=2)
            
                logger.info(f"✅ 完整schema已保存到: {output_file}")
            except Exception as e:
                logger.error(f"❌ 保存schema失败: {e}")
        except Exception as e:
            logger.error(f"❌ 从ready schema创建集合失败: {e}")
            return ""
    
    def _create_collections_from_llm_system(self, classification_system: Dict[str, Any], existing_collections: Dict[str, str], dry_run: bool = False) -> Dict[str, str]:
        """从LLM生成的分类体系创建集合"""
        main_categories = classification_system.get('main_categories', {})
        
        if not main_categories:
            logger.error("❌ 分类体系中没有主分类")
            return {}
        
        # 获取现有集合的名称和key映射
        existing_names = {name: key for key, name in existing_collections.items()}
        existing_keys = {name: key for key, name in existing_collections.items()}
        
        logger.info(f"📊 开始创建集合:")
        logger.info(f"   主分类数: {len(main_categories)}")
        logger.info(f"   现有集合: {len(existing_collections)}")
        
        # 创建主分类集合
        logger.info(f"🎯 创建主分类集合:")
        main_collections = {}
        created_main = 0
        all_collection_keys = {}

        for category_code, category_data in main_categories.items():
            category_name = category_data["name"]
            
            
            
            if dry_run:
                logger.info(f"🔍 [干运行] 将创建主分类: {category_name}")
            else:
                collection_key = self._create_collection(category_name, category_data.get("description", ""))
                if collection_key:
                    main_collections[category_code] = collection_key
                    created_main += 1
                    logger.info(f"✅ 创建主分类: {category_name} (key: {collection_key})")
                    existing_names[category_name] = collection_key
                    category_data['collection_key'] = collection_key
                    category_data['collection_key'] = collection_key
                else:
                    logger.error(f"❌ 创建主分类失败: {category_name}")
        
        logger.info(f"📊 主分类创建完成:")
        logger.info(f"   创建成功: {created_main} 个")
        logger.info(f"   跳过已存在: {len(main_categories) - created_main} 个")
        
        # 创建子分类集合
        logger.info(f"🎯 创建子分类集合:")
        total_subcategories = 0
        created_subcategories = 0
        
        for category_code, category_data in main_categories.items():
            subcategories = category_data.get("subcategories", {})
            total_subcategories += len(subcategories)
            
            if not subcategories:
                continue
            
            parent_name = category_data["name"]
            logger.info(f"📂 为 {parent_name} 创建子分类:")
            
            for sub_cat_code, sub_cat_info in subcategories.items():
                sub_name = sub_cat_info.get("name", "")
                
                
                
                if dry_run:
                    logger.info(f"🔍 [干运行] 将创建子分类: {sub_name} (父分类: {parent_name})")
                else:
                    parent_key = main_collections.get(category_code)
                    if parent_key:
                        collection_key = self._create_collection(sub_name, parent_key=parent_key)
                        if collection_key:
                            all_collection_keys[sub_cat_code] = collection_key
                            created_subcategories += 1
                            sub_cat_info['collection_key'] = collection_key
                            sub_cat_info['collection_key'] = collection_key
                    else:
                        logger.error(f"❌ 无法创建子分类 {sub_name} - 父分类 {parent_name} 不存在")
        
        logger.info(f"📊 创建完成统计:")
        logger.info(f"   主分类总数: {len(main_categories)}")
        logger.info(f"   子分类总数: {total_subcategories}")
        logger.info(f"   实际创建子分类: {created_subcategories}")
        
        # 保存更新后的schema（包含collection_key）
        if not dry_run:
            # 创建带collection_key的schema文件
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            updated_schema_file = f"data/schema_with_collection_keys_{timestamp}.json"
            
            # 构建完整的schema
            complete_schema = {
                "metadata": {
                    "created_at": datetime.now().isoformat(),
                    "status": "collections_created",
                    "total_collections_created": len(main_collections) + created_subcategories
                },
                "classification_schema": classification_system,
                "collection_mapping": main_collections
            }
            
            with open(updated_schema_file, 'w', encoding='utf-8') as f:
                json.dump(complete_schema, f, ensure_ascii=False, indent=2)
            
            logger.info(f"✅ 所有集合创建完成！")
            logger.info(f"📁 更新后的schema已保存到: {updated_schema_file}")
            return updated_schema_file
        else:
            logger.info(f"🔍 干运行完成，未实际创建任何集合")
            return None
    
    def _get_existing_collections(self) -> Dict[str, str]:
        """获取现有集合"""
        try:
            # 使用配置系统
            zotero_config = get_zotero_config()
            
            url = f"https://api.zotero.org/users/{zotero_config.user_id}/collections"
            headers = zotero_config.headers
            
            response = requests.get(url, headers=headers)
            response.raise_for_status()
            
            collections = response.json()
            collection_dict = {}
            
            for collection in collections:
                key = collection.get('key')
                name = collection.get('data', {}).get('name', '')
                if key and name:
                    collection_dict[key] = name
            
            logger.info(f"✅ 获取到 {len(collection_dict)} 个现有集合")
            return collection_dict
            
        except Exception as e:
            logger.error(f"❌ 获取集合异常: {e}")
            return {}
    
    def _create_collection(self, name: str, description: str = "", parent_key: str = None) -> str:
        """创建集合"""
        try:
            # 使用配置系统
            zotero_config = get_zotero_config()
            
            url = f"https://api.zotero.org/users/{zotero_config.user_id}/collections"
            headers = zotero_config.headers
            
            # 构建集合数据
            collection_data = [{
                "name": name
            }]
            
            # 如果有父集合，添加到父集合下
            if parent_key:
                collection_data[0]["parentCollection"] = parent_key
            
            response = requests.post(url, headers=headers, json=collection_data)
            response.raise_for_status()
            
            result = response.json()
            if result.get('successful') and '0' in result['successful']:
                collection_key = result['successful']['0']['key']
            else:
                logger.error(f"❌ 创建集合失败: {result}")
                collection_key = None
            
            if collection_key:
                logger.info(f"✅ 创建集合成功: {name} (key: {collection_key})")
                return collection_key
            else:
                logger.error(f"❌ 创建集合失败: 未获取到key")
                return ""
        except Exception as e:
            logger.error(f"❌ 创建集合失败: {e}")
            return ""
    
    def save_collection_mapping(self, output_file: str = None) -> str:
        """保存集合映射"""
        if output_file is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = f"data/collection_mapping_{timestamp}.json"
        
        try:
            os.makedirs(os.path.dirname(output_file), exist_ok=True)
            with open(output_file, 'w', encoding='utf-8') as f:
                json.dump(self.collection_keys, f, ensure_ascii=False, indent=2)
            
            logger.info(f"✅ 集合映射已保存到: {output_file}")
            return output_file
        except Exception as e:
            logger.error(f"❌ 保存集合映射失败: {e}")
            return ""


def main():
    """主函数"""
    parser = argparse.ArgumentParser(
        description="002 - 基于文献信息生成分类Schema，并可选地在Zotero中创建对应的集合结构",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
使用示例:
  # 第一步：生成分类体系文件（安全操作）
  python 002_generate_schema_and_create_collections.py --generate-schema --input data/literature_info.xlsx --max-items 100
  
  # 第二步：创建Zotero集合（危险操作）
  python 002_generate_schema_and_create_collections.py --create-collections --schema data/classification_schema_ready.json
  
  # 使用干运行模式预览第二步
  python 002_generate_schema_and_create_collections.py --create-collections --schema data/classification_schema_ready.json --dry-run

  # 测试模式
  python 002_generate_schema_and_create_collections.py --test --input data/literature_info.xlsx

注意事项:
  - 需要配置LLM API环境变量
  - 需要配置Zotero API环境变量
  - 建议先使用--test或--dry-run模式测试
  - 创建集合操作会永久修改Zotero库，请谨慎操作
        """
    )
    
    # 创建互斥组
    mode_group = parser.add_mutually_exclusive_group(required=True)
    mode_group.add_argument('--test', action='store_true', help='测试模式（使用较少数据）')
    mode_group.add_argument('--generate-schema', action='store_true', help='第一步：生成分类体系文件（安全操作）')
    mode_group.add_argument('--create-collections', action='store_true', help='第二步：创建Zotero集合（危险操作）')
    
    # 文件路径参数
    parser.add_argument('--input', type=str, help='文献数据文件路径（Excel格式）')
    parser.add_argument('--schema', type=str, help='分类schema文件路径（JSON格式）')
    
    # 可选参数
    parser.add_argument('--max-items', type=int, help='最大处理文献数量（默认使用所有文献）')
    parser.add_argument('--dry-run', action='store_true', help='干运行模式，只显示计划，不实际创建')
    
    args = parser.parse_args()
    
    # 根据模式验证必需参数
    if args.test or args.generate_schema:
        if not args.input:
            parser.error("--test 和 --generate-schema 模式需要指定 --input 参数")
        if not os.path.exists(args.input):
            parser.error(f"输入文件不存在: {args.input}")
    
    if args.create_collections:
        if not args.schema:
            parser.error("--create-collections 模式需要指定 --schema 参数")
        if not os.path.exists(args.schema):
            parser.error(f"Schema文件不存在: {args.schema}")
    
    # 创建管理器（根据模式决定初始化）
    if args.create_collections:
        # 创建集合模式需要Zotero客户端
        manager = SchemaBasedCollectionManager(init_llm=False, init_zotero=True)
    else:
        # 其他模式需要LLM客户端
        manager = SchemaBasedCollectionManager(init_llm=True, init_zotero=False)
    
    # 根据模式执行相应操作
    if args.test:
        print(f"\n🧪 测试模式：使用 {args.max_items or get_default_test_items()} 篇文献生成分类体系...")
        
        classification_system = manager.generate_collections_from_literature(
            literature_file=args.input, 
            max_items=args.max_items or get_default_test_items(), 
            dry_run=False,
            return_schema_only=True
        )
        
        if classification_system:
            schema_file = manager.save_llm_generated_schema(classification_system)
            if schema_file:
                print(f"\n✅ 测试完成！分类体系已保存到: {schema_file}")
                return 0
            else:
                print("❌ 保存分类体系失败")
                return 1
        else:
            print("❌ 生成分类体系失败")
            return 1
            
    elif args.generate_schema:
        print(f"\n📝 第一步：生成分类体系文件（安全操作）...")
        
        # 生成并修复schema
        classification_system = manager.generate_collections_from_literature(
            literature_file=args.input, 
            max_items=args.max_items, 
            dry_run=False,
            return_schema_only=True
        )
        
        if classification_system:
            # 保存为ready状态的文件
            ready_schema_file = manager.save_ready_schema(classification_system)
            if ready_schema_file:
                excel_file = ready_schema_file.replace('.json', '.xlsx')
                print(f"\n✅ 第一步完成！分类体系已保存。")
                print(f"   - JSON: {ready_schema_file}")
                print(f"   - Excel: {excel_file}")
                print("\n📋 下一步操作:")
                print(f"  1. 检查生成的文件: {excel_file} 和 {ready_schema_file}")
                print(f"  2. 确认分类体系合理后，运行:")
                print(f"     python 003_convert_schema_format.py --new-to-old --input {ready_schema_file}")
                return 0
            else:
                print("❌ 保存分类体系失败")
                return 1
        else:
            print("❌ 生成分类体系失败")
            return 1
        
    elif args.create_collections:
        print(f"\n🏗️  第二步：创建Zotero集合...")
        
        # 显示操作摘要
        summary = manager.get_operation_summary(args.schema)
        print(f"\n📊 操作摘要:")
        print(f"  主分类数: {summary['main_categories']}")
        print(f"  子分类数: {summary['sub_categories']}")
        print(f"  总分类数: {summary['total_categories']}")
        
        if not args.dry_run:
            confirm = input(f"\n⚠️  这是一个危险操作，将永久修改您的Zotero库。确认要创建这些集合吗？(y/N): ").strip().lower()
            if confirm != 'y':
                print("操作已取消")
                return 0
    
        # 创建集合
        result_file = manager.create_collections_from_ready_schema(args.schema, args.dry_run)
        if result_file or args.dry_run:
            if args.dry_run:
                print("\n✅ 干运行完成，未对Zotero进行任何修改。")
            else:
                print(f"✅ 第二步完成！集合已创建，完整schema已保存到: {result_file}")
                print("📋 下一步操作:")
                print(f"  1. 检查Zotero中的集合结构")
                print(f"  2. 运行: python 004_reclassify_with_new_schema.py --plan --schema {result_file}")
            return 0
        else:
            print("❌ 创建集合失败")
            return 1


if __name__ == "__main__":
    sys.exit(main())