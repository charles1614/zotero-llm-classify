#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
003 - Reclassify Literature with New Schema
使用LLM和新的分类体系对文献进行智能分类

主要功能：
1. 加载新的分类体系schema
2. 使用LLM对文献进行智能分类
3. 支持批量处理和测试模式
4. 生成分类计划并应用到Zotero

注意：此脚本专注于文献分类，集合创建在005脚本中实现
"""

import os
import sys
import json
import pandas as pd
import time
import requests
import argparse
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
import logging
from tqdm import tqdm
import multiprocessing as mp
from concurrent.futures import ThreadPoolExecutor, as_completed

# 导入配置系统
from config import (
    get_llm_config, get_zotero_config, get_config,
    get_default_batch_size, get_default_test_items, get_default_max_items,
    get_max_tokens_limit, get_default_output_tokens,
    get_title_preview_length, get_description_preview_length
)

# 设置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# 导入自定义模块
from llm_client import LLMClient

class NewSchemaLiteratureClassifier:
    """基于新schema的文献分类器"""
    
    def __init__(self):
        """初始化分类器"""
        self.data_dir = Path("data")
        
        # 初始化LLM客户端
        self.llm_client = self._init_llm_client()
        
        # 统计信息
        self.total_items = 0
        self.processed_items = 0
        self.successful_classifications = 0
        self.failed_classifications = 0
        
    def _init_llm_client(self) -> Optional[LLMClient]:
        """初始化LLM客户端"""
        try:
            # 使用新的配置系统
            return LLMClient()
        except Exception as e:
            logger.error(f"❌ 初始化LLM客户端失败: {e}")
            return None
    
    def _load_schema(self, schema_file: str) -> Dict[str, Any]:
        """加载schema文件"""
        try:
            with open(schema_file, 'r', encoding='utf-8') as f:
                schema = json.load(f)
            logger.info(f"✅ 成功加载schema: {schema_file}")
            return schema
        except Exception as e:
            logger.error(f"❌ 加载schema失败: {e}")
            sys.exit(1)
    
    def _build_collection_mapping(self, schema: Dict[str, Any]) -> Dict[str, Dict[str, str]]:
        """构建集合映射（只包含可分配的子分类）"""
        collection_mapping = {}
        
        # 始终从classification_schema构建，以确保只包含子分类
        classification_schema = schema.get('classification_schema', {})
        if not classification_schema:
            logger.warning("⚠️ schema中未找到 'classification_schema'，无法构建集合映射。")
            return {}

        main_categories = classification_schema.get('main_categories', {})
    
        for cat_code, cat_info in main_categories.items():
            # 不直接添加主分类
            
            # 处理subcategories
            subcategories = cat_info.get('subcategories', [])
            if isinstance(subcategories, list):
                for sub_info in subcategories:
                    sub_name = sub_info.get('name', '')
                    collection_key = sub_info.get('collection_key', '')
                    sub_description = sub_info.get('description', '')
                    if sub_name and collection_key:
                        collection_mapping[collection_key] = {
                            'name': sub_name,
                            'description': sub_description
                        }
            elif isinstance(subcategories, dict):
                for sub_code, sub_info in subcategories.items():
                    sub_name = sub_info.get('name', '')
                    sub_description = sub_info.get('description', '')
                    if sub_name:
                        collection_mapping[sub_code] = {
                            'name': sub_name,
                            'description': sub_description
                        }
        
        logger.info(f"✅ 构建仅包含子分类的集合映射: {len(collection_mapping)} 个集合")
        return collection_mapping
    
    def _load_literature_data(self, literature_file: str) -> List[Dict[str, Any]]:
        """加载文献数据"""
        try:
            # 支持Excel和JSON格式
            if literature_file.endswith('.xlsx'):
                df = pd.read_excel(literature_file)
                literature_data = df.to_dict('records')
            elif literature_file.endswith('.json'):
                with open(literature_file, 'r', encoding='utf-8') as f:
                    literature_data = json.load(f)
            else:
                logger.error(f"❌ 不支持的文件格式: {literature_file}")
                return []
            
            logger.info(f"✅ 成功加载文献数据: {len(literature_data)} 篇文献")
            return literature_data
            
        except Exception as e:
            logger.error(f"❌ 加载文献数据失败: {e}")
            return []
    
    def _prepare_classification_prompt(self, item: Dict[str, Any], collection_mapping: Dict[str, Dict[str, str]]) -> str:
        """准备分类提示词"""
        title = str(item.get('title', '')).strip()
        abstract = str(item.get('abstract', '')).strip()
        
        # 构建集合列表
        collection_list = []
        for code, info in collection_mapping.items():
            name = info.get('name', '')
            description = info.get('description', '')
            collection_list.append(f"- {code}: {name} - {description}")
        
        collections_text = "\n".join(collection_list)
        
        prompt = f"""请根据以下文献信息，从给定的集合中选择最合适的分类。

文献信息：
标题：{title}
摘要：{abstract}

可用集合：
{collections_text}

请严格按照以下JSON格式返回分类结果：
{{
    "recommended_collections": ["collection_code1", "collection_code2"],
    "reasoning": "分类理由说明"
}}

要求：
1. recommended_collections: 最多推荐5个集合，按优先级排序
2. 只使用上述集合代码，不要创建新的分类
3. 确保集合代码完全匹配
4. reasoning: 简要说明分类理由
5. 如果文献与任何集合都不匹配，返回空数组

请只返回JSON格式，不要包含其他内容。"""
        
        return prompt
    
    def _parse_classification_response(self, response: str) -> Dict[str, Any]:
        """解析分类响应"""
        try:
            # 尝试提取JSON部分
            start_idx = response.find('{')
            end_idx = response.rfind('}') + 1
            
            if start_idx == -1 or end_idx == 0:
                return {"recommended_collections": [], "reasoning": "无法解析响应"}
            
            json_str = response[start_idx:end_idx]
            result = json.loads(json_str)
            
            # 验证响应格式
            if 'recommended_collections' not in result:
                result['recommended_collections'] = []
            if 'reasoning' not in result:
                result['reasoning'] = "未提供分类理由"
            
            return result
            
        except json.JSONDecodeError as e:
            logger.warning(f"JSON解析失败: {e}")
            return {"recommended_collections": [], "reasoning": f"JSON解析失败: {e}"}
        except Exception as e:
            logger.warning(f"解析响应失败: {e}")
            return {"recommended_collections": [], "reasoning": f"解析失败: {e}"}
    
    def _prepare_batch_classification_prompt(self, items: List[Dict[str, Any]], collection_mapping: Dict[str, Dict[str, str]]) -> str:
        """准备批量分类提示词"""
        # 构建集合列表
        collection_list = []
        for code, info in collection_mapping.items():
            name = info.get('name', '')
            description = info.get('description', '')
            collection_list.append(f"- {code}: {name} - {description}")
        
        collections_text = "\n".join(collection_list)
        
        # 构建文献列表
        items_text = ""
        for i, item in enumerate(items, 1):
            title = str(item.get('title', '')).strip()
            abstract = str(item.get('abstract', '')).strip()
            item_key = item.get('item_key', '')
            
            items_text += f"""
文献 {i} (ID: {item_key}):
标题：{title}
摘要：{abstract}
"""
        
        prompt = f"""
# ROLE: You are a professional AI literature classification engine.

# CORE TASK: Your primary task is to accurately assign each document from a given list (`items_text`) to one or more relevant categories from a predefined, flat list of collections (`collections_text`).

---

### Input Data

1.  **Available Collections (`{collections_text}`)**:
    * **Format**: A flat JSON list of available classification categories. Each category object contains a `collection_key`,`name`,`description`.
    * **Example**: `- 9KGVHHUD: Foundation Models - Large-scale models pre-trained on vast data, serving as a base for various downstream tasks, such as GPT-3, Llama 3, and ERNIE 4.5.\n- T6PHSH3J: Large Language Models (LLMs) - Models specifically designed for understanding, generating, and processing natural language, including architectures, training methodologies, and few-shot learning capabilities.`

2.  **Items to Classify (`{items_text}`)**:
    * **Format**: A JSON list of documents, where each document has a unique `literature`, `title` and `abstract`.

---

### Core Requirements

1.  **High-Confidence Principle: DO NOT FORCE CLASSIFICATION.** You must first carefully read and internalize the `description` of each available category. Only recommend a category if the document's core topic **clearly and strongly aligns** with the category's description. Avoid all weak, speculative, or overly broad matches.
2.  **Semantic Matching**: Perform a precise semantic match between a document's primary research contribution and the category descriptions. Focus on the core problem being solved, not just shared keywords.
3.  **Ranking and Limits**: If confident matches are found, recommend **1 to 5** of the most relevant categories. The results in the `recommended_collections` array MUST be sorted by relevance, from **highest to lowest**.
4.  **Code Integrity**: The `recommended_collections` array MUST ONLY contain the `code` values from the provided `collections_text` list. Ensure the codes match exactly.
5.  **Provide Reasoning**: In the `reasoning` field, provide a brief, one-sentence explanation that justifies the high-confidence match by linking the document's specific contribution to the category's definition.
6.  **No-Match Handling**: Following the High-Confidence Principle, if a document does not have a strong and clear match with any category, the `recommended_collections` field must be an **empty array `[]`**.
7.  **Maintain Order**: The order of the documents in your final output must be **exactly the same** as the order in the input `items_text`.

---

### Output Format

You MUST strictly adhere to the following JSON structure. Do not include any text, notes, or explanations outside of this JSON object.

```json
{{
    "classifications": [
        {{
            "item_key": "item_key_of_document_1",
            "recommended_collections": ["collection_code1", "collection_code2"],
            "reasoning": "A brief explanation of why the document was assigned to these collections."
        }},
        {{
            "item_key": "item_key_of_document_2",
            "recommended_collections": ["collection_code3"],
            "reasoning": "A brief explanation of why the document was assigned to this collection."
        }}
    ]
}}
```
"""

        return prompt
    
    def _parse_batch_classification_response(self, response: str, items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """解析批量分类响应"""
        try:
            # 尝试提取JSON部分
            start_idx = response.find('{')
            end_idx = response.rfind('}') + 1
            
            if start_idx == -1 or end_idx == 0:
                # 如果无法解析，为所有文献返回失败结果
                return [{
                    'item_key': item.get('item_key', ''),
                    'title': item.get('title', ''),
                    'classification_success': False,
                    'recommended_collections': [],
                    'reasoning': '无法解析响应',
                    'error_message': '响应格式错误'
                } for item in items]
            
            json_str = response[start_idx:end_idx]
            result = json.loads(json_str)
            
            # 验证响应格式
            if 'classifications' not in result:
                # 如果格式不正确，为所有文献返回失败结果
                return [{
                    'item_key': item.get('item_key', ''),
                    'title': item.get('title', ''),
                    'classification_success': False,
                    'recommended_collections': [],
                    'reasoning': '响应格式错误',
                    'error_message': '缺少classifications字段'
                } for item in items]
            
            classifications = result['classifications']
            results = []
            
            # 为每个文献创建结果
            for i, item in enumerate(items):
                item_key = item.get('item_key', '')
                
                # 查找对应的分类结果
                classification = None
                for cls in classifications:
                    if cls.get('item_key') == item_key:
                        classification = cls
                        break
                
                if classification and 'recommended_collections' in classification:
                    results.append({
                        'item_key': item_key,
                        'title': item.get('title', ''),
                        'classification_success': len(classification['recommended_collections']) > 0,
                        'recommended_collections': classification.get('recommended_collections', []),
                        'reasoning': classification.get('reasoning', ''),
                        'error_message': '' if len(classification['recommended_collections']) > 0 else '未找到合适的分类'
                    })
                else:
                    results.append({
                        'item_key': item_key,
                        'title': item.get('title', ''),
                        'classification_success': False,
                        'recommended_collections': [],
                        'reasoning': '未找到对应的分类结果',
                        'error_message': '响应中缺少该文献的分类信息'
                    })
            
            return results
            
        except Exception as e:
            logger.error(f"解析批量分类响应失败: {e}")
            # 返回失败结果
            return [{
                'item_key': item.get('item_key', ''),
                'title': item.get('title', ''),
                'classification_success': False,
                'recommended_collections': [],
                'reasoning': f'解析失败: {str(e)}',
                'error_message': '响应解析异常'
            } for item in items]

    def _classify_batch(self, items: List[Dict[str, Any]], collection_mapping: Dict[str, Dict[str, str]]) -> List[Dict[str, Any]]:
        """批量分类文献"""
        if not self.llm_client:
            logger.error("❌ LLM客户端未初始化")
            return [{
                'item_key': item.get('item_key', ''),
                'title': item.get('title', ''),
                'classification_success': False,
                'recommended_collections': [],
                'reasoning': '',
                'error_message': 'LLM客户端未初始化'
            } for item in items]
        
        try:
            # 准备批量分类提示词
            prompt = self._prepare_batch_classification_prompt(items, collection_mapping)
            
            # 调用LLM API
            response = self.llm_client.generate_text(prompt)
            
            if not response:
                logger.error("❌ LLM API返回空响应")
                return [{
                    'item_key': item.get('item_key', ''),
                    'title': item.get('title', ''),
                    'classification_success': False,
                    'recommended_collections': [],
                    'reasoning': '',
                    'error_message': 'LLM API返回空响应'
                } for item in items]
            
            # 解析响应
            results = self._parse_batch_classification_response(response, items)
            
            # 统计成功数量
            successful = sum(1 for r in results if r['classification_success'])
            logger.info(f"📊 批量分类完成: {len(items)} 篇, 成功: {successful} 篇")
            
            return results
    
        except Exception as e:
            logger.error(f"批量分类失败: {e}")
            return [{
                'item_key': item.get('item_key', ''),
                'title': item.get('title', ''),
                'classification_success': False,
                'recommended_collections': [],
                'reasoning': '',
                'error_message': str(e)
            } for item in items]
    
    def classify_literature(self, schema_file: str, literature_file: str, max_items: int = None, batch_size: int = None) -> str:
        """对文献进行分类"""
        # 加载schema和集合映射
        schema = self._load_schema(schema_file)
        collection_mapping = self._build_collection_mapping(schema)
        
        # 加载文献数据
        literature_data = self._load_literature_data(literature_file)
        if not literature_data:
            logger.error("❌ 没有可分类的文献数据")
            return ""
        
        # 限制处理数量
        if max_items:
            literature_data = literature_data[:max_items]
        
        self.total_items = len(literature_data)
        logger.info(f"📊 开始分类 {self.total_items} 篇文献")
        
        # 批量处理
        batch_size = batch_size or get_default_batch_size()
        results = []
        
        for i in range(0, len(literature_data), batch_size):
            batch = literature_data[i:i + batch_size]
            logger.info(f"📦 处理批次 {i//batch_size + 1}/{(len(literature_data) + batch_size - 1)//batch_size}")
            
            batch_results = self._classify_batch(batch, collection_mapping)
            results.extend(batch_results)
            
            # 统计进度
            successful = sum(1 for r in batch_results if r['classification_success'])
            logger.info(f"✅ 批次完成: {len(batch_results)} 篇, 成功: {successful} 篇")
        
        # 保存结果
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = f"data/classification_plan_{timestamp}.json"
        excel_file = f"data/classification_plan_{timestamp}.xlsx"
        
        output_data = {
            'metadata': {
                'generated_at': datetime.now().isoformat(),
                'schema_file': schema_file,
                'literature_file': literature_file,
                'total_items': len(results),
                'successful_classifications': sum(1 for r in results if r['classification_success']),
                'failed_classifications': sum(1 for r in results if not r['classification_success'])
            },
            'classifications': results
        }
        
        try:
            os.makedirs(os.path.dirname(output_file), exist_ok=True)
            
            # 保存JSON文件
            with open(output_file, 'w', encoding='utf-8') as f:
                json.dump(output_data, f, ensure_ascii=False, indent=2)
        
            logger.info(f"✅ 分类计划已保存到: {output_file}")
            
            # 生成Excel文件
            self._save_excel_report(results, excel_file, collection_mapping, literature_file)
            logger.info(f"✅ Excel报告已保存到: {excel_file}")
            
            return output_file
            
        except Exception as e:
            logger.error(f"❌ 保存分类计划失败: {e}")
            return ""
        
    def _save_excel_report(self, results: List[Dict[str, Any]], excel_file: str, collection_mapping: Dict[str, Dict[str, str]], literature_file: str) -> None:
        """保存Excel格式的分类报告"""
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            # 读取原始文献数据
            original_data = self._load_literature_data(literature_file)
            
            # 创建结果映射字典
            results_dict = {result['item_key']: result for result in results}
            
            # 为每个原始文献添加新的分类结果
            excel_data = []
            for item in original_data:
                item_key = item.get('item_key', '')
                result = results_dict.get(item_key, {})
                
                # 将推荐集合列表转换为可读的文本
                recommended_collections = result.get('recommended_collections', [])
                collections_text = ""
                if recommended_collections:
                    collection_names = []
                    for code in recommended_collections:
                        info = collection_mapping.get(code, {})
                        name = info.get('name', code) if isinstance(info, dict) else code
                        collection_names.append(f"{code}: {name}")
                    collections_text = "; ".join(collection_names)
                
                # 创建行数据，保持原有格式
                row_data = {
                    'item_key': item.get('item_key', ''),
                    'title': item.get('title', ''),
                    'item_type': item.get('item_type', ''),
                    'authors': item.get('authors', ''),
                    'publication_title': item.get('publication_title', ''),
                    'conference_name': item.get('conference_name', ''),
                    'date': item.get('date', ''),
                    'doi': item.get('doi', ''),
                    'abstract': item.get('abstract', ''),
                    'tags': item.get('tags', ''),
                    'url': item.get('url', ''),
                    'language': item.get('language', ''),
                    'pages': item.get('pages', ''),
                    'volume': item.get('volume', ''),
                    'issue': item.get('issue', ''),
                    'publisher': item.get('publisher', ''),
                    'place': item.get('place', ''),
                    'edition': item.get('edition', ''),
                    'series': item.get('series', ''),
                    'isbn': item.get('isbn', ''),
                    'issn': item.get('issn', ''),
                    'call_number': item.get('call_number', ''),
                    'access_date': item.get('access_date', ''),
                    'rights': item.get('rights', ''),
                    'extra': item.get('extra', ''),
                    'collections': item.get('collections', ''),
                    'collections_keys': item.get('collections_keys', ''),
                    'collections_count': item.get('collections_count', 0),
                    'notes': item.get('notes', ''),
                    'attachments': item.get('attachments', ''),
                    'attachments_count': item.get('attachments_count', 0),
                    'related_items': item.get('related_items', ''),
                    'related_items_count': item.get('related_items_count', 0),
                    'created_date': item.get('created_date', ''),
                    'modified_date': item.get('modified_date', ''),
                    'last_modified_by': item.get('last_modified_by', ''),
                    'version': item.get('version', ''),
                    # 新增分类结果列
                    'new_classification_success': result.get('classification_success', False),
                    'new_recommended_collection_keys': '; '.join(recommended_collections) if recommended_collections else '',
                    'new_recommended_collections': collections_text,
                    'new_recommended_count': len(recommended_collections),
                    'new_analysis': result.get('reasoning', ''),
                    'new_error_message': result.get('error_message', ''),
                    'new_worker_id': '006_reclassify_with_new_schema',
                    'new_response': '',  # 可以添加原始响应
                    'new_classification_timestamp': datetime.now().strftime("%Y%m%d_%H%M%S")
                }
                excel_data.append(row_data)
        
            # 创建DataFrame
            df = pd.DataFrame(excel_data)
        
            # 创建Excel文件
            with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
                # 写入主数据表
                df.to_excel(writer, sheet_name='分类结果', index=False)
                
                # 获取工作表对象
                worksheet = writer.sheets['分类结果']
                
                # 设置列宽
                column_widths = {
                    'A': 15,  # item_key
                    'B': 50,  # title
                    'C': 15,  # item_type
                    'D': 30,  # authors
                    'E': 40,  # publication_title
                    'F': 30,  # conference_name
                    'G': 15,  # date
                    'H': 25,  # doi
                    'I': 60,  # abstract
                    'J': 30,  # tags
                    'K': 40,  # url
                    'L': 10,  # language
                    'M': 10,  # pages
                    'N': 10,  # volume
                    'O': 10,  # issue
                    'P': 25,  # publisher
                    'Q': 20,  # place
                    'R': 15,  # edition
                    'S': 20,  # series
                    'T': 20,  # isbn
                    'U': 15,  # issn
                    'V': 20,  # call_number
                    'W': 15,  # access_date
                    'X': 20,  # rights
                    'Y': 30,  # extra
                    'Z': 40,  # collections
                    'AA': 40, # collections_keys
                    'AB': 15, # collections_count
                    'AC': 40, # notes
                    'AD': 40, # attachments
                    'AE': 15, # attachments_count
                    'AF': 40, # related_items
                    'AG': 15, # related_items_count
                    'AH': 20, # created_date
                    'AI': 20, # modified_date
                    'AJ': 20, # last_modified_by
                    'AK': 10, # version
                    'AL': 20, # new_classification_success
                    'AM': 40, # new_recommended_collection_keys
                    'AN': 60, # new_recommended_collections
                    'AO': 15, # new_recommended_count
                    'AP': 60, # new_analysis
                    'AQ': 30, # new_error_message
                    'AR': 25, # new_worker_id
                    'AS': 40, # new_response
                    'AT': 20  # new_classification_timestamp
                }
                
                for col, width in column_widths.items():
                    worksheet.column_dimensions[col].width = width
                
                # 设置标题行样式
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # 设置数据行样式
                success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # 浅绿色
                failure_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # 浅红色
                
                for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
                    # 根据分类成功状态设置背景色（只对新分类列设置背景色）
                    success_cell = row[37]  # new_classification_success列 (AL列，索引37)
                    if success_cell.value == True:
                        # 只对新分类相关的列设置绿色背景
                        for i in range(37, 46):  # AL到AT列
                            if i < len(row):
                                row[i].fill = success_fill
                    else:
                        # 只对新分类相关的列设置红色背景
                        for i in range(37, 46):  # AL到AT列
                            if i < len(row):
                                row[i].fill = failure_fill
                    
                    # 设置文本换行
                    for cell in row:
                        cell.alignment = Alignment(wrap_text=True, vertical="top")
        
                # 创建统计信息表
                stats_data = {
                    '统计项目': [
                        '总文献数',
                        '分类成功数',
                        '分类失败数',
                        '成功率',
                        '生成时间'
                    ],
                    '数值': [
                        len(results),
                        sum(1 for r in results if r['classification_success']),
                        sum(1 for r in results if not r['classification_success']),
                        f"{sum(1 for r in results if r['classification_success']) / len(results) * 100:.1f}%" if results else "0%",
                        datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    ]
                }
                
                stats_df = pd.DataFrame(stats_data)
                stats_df.to_excel(writer, sheet_name='统计信息', index=False)
                
                # 设置统计表样式
                stats_worksheet = writer.sheets['统计信息']
                stats_worksheet.column_dimensions['A'].width = 15
                stats_worksheet.column_dimensions['B'].width = 20
                
                for cell in stats_worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # 创建集合映射表
                mapping_data = []
                for code, info in collection_mapping.items():
                    name = info.get('name', '') if isinstance(info, dict) else str(info)
                    description = info.get('description', '') if isinstance(info, dict) else ''
                    mapping_data.append({
                        '集合代码': code,
                        '集合名称': name,
                        '集合描述': description
                    })
                
                mapping_df = pd.DataFrame(mapping_data)
                mapping_df.to_excel(writer, sheet_name='集合映射', index=False)
                
                # 设置映射表样式
                mapping_worksheet = writer.sheets['集合映射']
                mapping_worksheet.column_dimensions['A'].width = 20
                mapping_worksheet.column_dimensions['B'].width = 40
                mapping_worksheet.column_dimensions['C'].width = 60
                
                for cell in mapping_worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
        except ImportError as e:
            logger.warning(f"⚠️ 无法生成Excel文件，缺少依赖: {e}")
            logger.info("请安装: pip install pandas openpyxl")
        except Exception as e:
            logger.error(f"❌ 生成Excel文件失败: {e}")
        

def main():
    """主函数"""
    parser = argparse.ArgumentParser(
        description="003 - 使用LLM和新的分类Schema对文献进行智能分类，生成分类计划",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
使用示例:
  # 生成分类计划
  python 006_reclassify_with_new_schema.py --plan --schema data/schema_with_collection_keys.json --input data/literature_info.xlsx
  
  # 测试模式
  python 006_reclassify_with_new_schema.py --test --schema data/schema_with_collection_keys.json --input data/literature_info.xlsx
  
  # 指定批量大小
  python 006_reclassify_with_new_schema.py --plan --schema data/schema_with_collection_keys.json --input data/literature_info.xlsx --batch-size 25

注意事项:
  - 需要配置LLM API环境变量
  - 建议先使用--test模式测试
  - 分类结果需要手动应用到Zotero
        """
    )
    
    # 创建互斥组
    mode_group = parser.add_mutually_exclusive_group(required=True)
    mode_group.add_argument('--test', action='store_true', help='测试模式（使用较少数据）')
    mode_group.add_argument('--plan', action='store_true', help='生成分类计划')
    
    # 文件路径参数（强制要求）
    parser.add_argument('--schema', type=str, required=True, help='分类schema文件路径（JSON格式）')
    parser.add_argument('--input', type=str, required=True, help='文献数据文件路径（Excel或JSON格式）')
    
    # 可选参数
    parser.add_argument('--max-items', type=int, help='最大处理文献数量')
    parser.add_argument('--batch-size', type=int, help='批量处理大小')
    
    args = parser.parse_args()
    
    # 验证文件存在
    if not os.path.exists(args.schema):
        parser.error(f"Schema文件不存在: {args.schema}")
    if not os.path.exists(args.input):
        parser.error(f"输入文件不存在: {args.input}")
    
        # 创建分类器
    classifier = NewSchemaLiteratureClassifier()
        
    # 根据模式执行
    if args.test:
        print(f"\n🧪 测试模式：使用 {args.max_items or get_default_test_items()} 篇文献进行测试...")
        max_items = args.max_items or get_default_test_items()
    else:
        print(f"\n📝 生成分类计划...")
        max_items = args.max_items
    
    # 执行分类
    result_file = classifier.classify_literature(
        schema_file=args.schema,
        literature_file=args.input,
        max_items=max_items,
        batch_size=args.batch_size
    )
            
    if result_file:
        # 生成对应的Excel文件名
        excel_file = result_file.replace('.json', '.xlsx')
        print(f"\n✅ 分类完成！结果已保存到:")
        print(f"  📄 JSON格式: {result_file}")
        print(f"  📊 Excel格式: {excel_file}")
        print(f"\n💡 下一步操作:")
        print(f"  1. 查看Excel报告: {excel_file}")
        print(f"  2. 检查JSON数据: {result_file}")
        print(f"  3. 应用分类到Zotero:")
        print(f"     python 005_apply_classification_to_zotero.py --plan {result_file} --test")
        return 0
    else:
        print("❌ 分类失败")
        return 1
        

if __name__ == "__main__":
    sys.exit(main()) 